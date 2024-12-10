import zipfile
from typing import IO, Any, BinaryIO, Dict

import openpyxl
import pytest
from aws_lambda_typing.context import Context
from openpyxl.worksheet.worksheet import Worksheet

from src.functions.subrecipient_treasury_report_gen import SubrecipientLambdaPayload
from src.lib.output_template_comparator import CPFFileArchive
from src.lib.treasury_generation_common import OrganizationObj, PreferencesObj, UserObj

_SAMPLE_VALID_XLSM = "tests/data/sample_valid.xlsm"
_SAMPLE_VALID_XLSM_V2024_05_24 = "tests/data/sample_valid_V2024_05_24.xlsm"
_SAMPLE_TEMPLATE = "tests/data/sample_template.xlsx"
_SAMPLE_TEMPLATE_2 = "tests/data/sample_template_2.xlsx"
_SAMPLE_TEMPLATE_1A = "tests/data/treasury/sample_1A_input_pass.xlsm"
_SAMPLE_TEMPLATE_1B = "tests/data/treasury/sample_1B_input_pass.xlsm"
_SAMPLE_TEMPLATE_1C = "tests/data/treasury/sample_1C_input_pass.xlsm"
_SAMPLE_TREASURY_OUTPUT_1A_XLSM = "tests/data/treasury/sample_1A_output.xlsx"
_SAMPLE_TREASURY_OUTPUT_1B_XLSM = "tests/data/treasury/sample_1B_output.xlsx"
_SAMPLE_TREASURY_OUTPUT_1C_XLSM = "tests/data/treasury/sample_1C_output.xlsx"
_SAMPLE_SUBRECIPIENT_TEMPLATE_EMPTY = (
    "tests/data/subrecipient_treasury_template_empty.xlsx"
)


@pytest.fixture
def valid_file() -> BinaryIO:
    file_path = _SAMPLE_VALID_XLSM_V2024_05_24
    return open(file_path, "rb")


@pytest.fixture
def valid_workbook() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_VALID_XLSM_V2024_05_24)


@pytest.fixture
def valid_workbook_old_schema() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_VALID_XLSM)


@pytest.fixture
def valid_workbook_old_compatible_schema() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_VALID_XLSM)
    workbook["Logic"]["B1"] = "v:20240401"
    return workbook


@pytest.fixture
def valid_coversheet(
    valid_workbook: openpyxl.Workbook,
) -> Worksheet:
    return valid_workbook["Cover"]


@pytest.fixture
def valid_project_sheet(
    valid_workbook: openpyxl.Workbook,
) -> Worksheet:
    return valid_workbook["Project"]


@pytest.fixture
def valid_workbook_1A() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_1A)


@pytest.fixture
def valid_project_sheet_1A() -> Worksheet:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_1A)["Project"]


@pytest.fixture
def valid_workbook_1A_with_conflict() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_TEMPLATE_1A)
    valid_project_sheet_1A = workbook["Project"]
    valid_project_sheet_1A["C13"] = "updated project 1a test"
    return workbook


@pytest.fixture
def valid_second_workbook_1A_sheet() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_TEMPLATE_1A)
    valid_project_sheet_1A = workbook["Project"]
    valid_project_sheet_1A["C13"] = "test 2"
    valid_project_sheet_1A["D13"] = "44"
    valid_project_sheet_1A["E13"] = "345634563457"
    valid_project_sheet_1A["F13"] = "345345346"
    return workbook


@pytest.fixture
def valid_workbook_1B() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_1B)


@pytest.fixture
def valid_project_sheet_1B() -> Worksheet:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_1B)["Project"]


@pytest.fixture
def valid_workbook_1B_with_conflict() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_TEMPLATE_1B)
    valid_project_sheet_1B = workbook["Project"]
    valid_project_sheet_1B["C13"] = "updated project 1B test"
    return workbook


@pytest.fixture
def valid_second_workbook_1B_sheet() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_TEMPLATE_1B)
    valid_project_sheet_1B = workbook["Project"]
    valid_project_sheet_1B["C13"] = "test 2"
    valid_project_sheet_1B["D13"] = "44"
    valid_project_sheet_1B["E13"] = "345634563457"
    valid_project_sheet_1B["F13"] = "345345346"
    return workbook


@pytest.fixture
def valid_workbook_1C() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_1C)


@pytest.fixture
def valid_project_sheet_1C() -> Worksheet:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_1C)["Project"]


@pytest.fixture
def valid_workbook_1C_with_conflict() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_TEMPLATE_1C)
    valid_project_sheet_1C = workbook["Project"]
    valid_project_sheet_1C["C13"] = "updated project 1c test"
    return workbook


@pytest.fixture
def valid_second_workbook_1C_sheet() -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(_SAMPLE_TEMPLATE_1C)
    valid_project_sheet_1C = workbook["Project"]
    valid_project_sheet_1C["C13"] = "test 2"
    valid_project_sheet_1C["D13"] = "44"
    valid_project_sheet_1C["E13"] = "345634563457"
    valid_project_sheet_1C["F13"] = "345345346"
    return workbook


@pytest.fixture
def valid_subrecipientsheet(
    valid_workbook: openpyxl.Workbook,
) -> Worksheet:
    return valid_workbook["Subrecipients"]


@pytest.fixture
def invalid_cover_sheet(valid_coversheet: Worksheet) -> Worksheet:
    valid_coversheet["A2"] = "INVALID"
    return valid_coversheet


@pytest.fixture
def invalid_cover_sheet_missing_code(
    valid_coversheet: Worksheet,
) -> Worksheet:
    valid_coversheet["A2"] = ""
    return valid_coversheet


@pytest.fixture
def invalid_cover_sheet_empty_code(
    valid_coversheet: Worksheet,
) -> Worksheet:
    valid_coversheet["A2"] = "  "
    return valid_coversheet


@pytest.fixture
def invalid_cover_sheet_empty_desc(
    valid_coversheet: Worksheet,
) -> Worksheet:
    valid_coversheet["B2"] = "  "
    return valid_coversheet


@pytest.fixture
def invalid_project_sheet(valid_project_sheet: Worksheet) -> Worksheet:
    valid_project_sheet["D13"] = "X" * 21
    return valid_project_sheet


@pytest.fixture
def invalid_project_sheet_missing_field(
    valid_project_sheet: Worksheet,
) -> Worksheet:
    valid_project_sheet["D13"] = ""
    return valid_project_sheet


@pytest.fixture
def invalid_project_sheet_empty_field(
    valid_project_sheet: Worksheet,
) -> Worksheet:
    valid_project_sheet["D13"] = "    "
    return valid_project_sheet


@pytest.fixture
def invalid_project_sheet_unmatching_subrecipient_tin_field(
    valid_workbook: openpyxl.Workbook,
) -> openpyxl.Workbook:
    valid_workbook["Subrecipients"]["E13"] = "123123124"
    return valid_workbook


@pytest.fixture
def invalid_project_sheet_unmatching_subrecipient_uei_field(
    valid_workbook: openpyxl.Workbook,
) -> openpyxl.Workbook:
    valid_workbook["Subrecipients"]["F13"] = "123412341235"
    return valid_workbook


@pytest.fixture
def invalid_project_sheet_unmatching_subrecipient_tin_uei_field(
    valid_workbook: openpyxl.Workbook,
) -> openpyxl.Workbook:
    valid_workbook["Subrecipients"]["E13"] = "123123124"
    valid_workbook["Subrecipients"]["F13"] = "123412341235"
    return valid_workbook


@pytest.fixture
def invalid_subrecipient_sheet(
    valid_subrecipientsheet: Worksheet,
) -> Worksheet:
    valid_subrecipientsheet["E13"] = "INVALID"
    return valid_subrecipientsheet


@pytest.fixture
def valid_subrecipient_sheet_blank_optional_fields(
    valid_subrecipientsheet: Worksheet,
) -> Worksheet:
    valid_subrecipientsheet["K13"] = ""
    valid_subrecipientsheet["M13"] = ""
    valid_subrecipientsheet["N13"] = ""
    return valid_subrecipientsheet


@pytest.fixture
def sample_template() -> IO[bytes]:
    return open(_SAMPLE_TEMPLATE, "rb")


@pytest.fixture
def template_workbook() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE)


@pytest.fixture
def template_workbook_two() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TEMPLATE_2)


@pytest.fixture
def cpf_file_archive(sample_template: IO[bytes]) -> CPFFileArchive:
    cpf_archive_file = zipfile.ZipFile("_tmp_test.zip", "w")
    cpf_archive_file.writestr("2024-05-19/TestFile.xlsx", sample_template.read())
    return CPFFileArchive(cpf_archive_file)


@pytest.fixture
def cpf_file_archive_two(sample_template: IO[bytes]) -> CPFFileArchive:
    cpf_archive_file = zipfile.ZipFile("_tmp_test.zip", "w")
    cpf_archive_file.writestr("2024-05-19/TestFile.xlsx", sample_template.read())
    cpf_archive_file.writestr("2024-05-19/TestFile2.xlsx", sample_template.read())
    return CPFFileArchive(cpf_archive_file)


@pytest.fixture
def output_1A_template() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TREASURY_OUTPUT_1A_XLSM)


@pytest.fixture
def output_1B_template() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TREASURY_OUTPUT_1B_XLSM)


@pytest.fixture
def output_1C_template() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_TREASURY_OUTPUT_1C_XLSM)


@pytest.fixture
def valid_aws_typing_context() -> Context:
    valid_context = Context()
    valid_context.aws_request_id = "dummy_aws_request_id"
    valid_context.log_group_name = "dummy_log_group_name"
    valid_context.log_stream_name = "dummy_log_stream_name"
    valid_context.function_name = "dummy_function_name"
    valid_context.memory_limit_in_mb = "128"
    valid_context.function_version = "$LATEST"
    valid_context.invoked_function_arn = (
        "arn:aws:lambda:dummy-region:123456789012:function:dummy_function_name"
    )
    return valid_context


@pytest.fixture
def valid_subrecipients_json_content() -> Dict[str, Any]:
    return {
        "subrecipients": [
            {
                "id": 1,
                "name": "subrecipient 3",
                "ueiTinCombo": "345634563456_345345345",
                "status": "ACTIVE",
                "organizationId": 1,
                "createdAt": "2024-07-11T03:15:13.050Z",
                "updatedAt": "2024-07-11T03:15:13.050Z",
                "subrecipientUploads": [
                    {
                        "id": 1,
                        "subrecipientId": 1,
                        "uploadId": 1,
                        "rawSubrecipient": {
                            "Name": "subrecipient 3",
                            "EIN__c": "345345345",
                            "Zip__c": "23432",
                            "City__c": "hyattsville",
                            "Zip_4__c": "null",
                            "Address__c": "lkagfsdj",
                            "POC_Name__c": "figgy",
                            "Address_2__c": "null",
                            "Address_3__c": "null",
                            "POC_Phone_Number__c": "2342342343",
                            "POC_Email_Address__c": "test@test",
                            "State_Abbreviated__c": "MD",
                            "Recipient_Profile_ID__c": "333",
                            "Unique_Entity_Identifier__c": "345634563456",
                        },
                        "createdAt": "2024-07-11T03:15:13.054Z",
                        "updatedAt": "2024-07-15T20:06:44.379Z",
                        "version": "v:20240524",
                    }
                ],
            }
        ]
    }


@pytest.fixture
def sample_subrecipients_generation_event() -> Dict[str, Any]:
    return {
        "organization": {
            "id": 12,
            "preferences": {"current_reporting_period_id": 34},
        },
        "user": {"id": 56, "email": "a@example.com"},
        "outputTemplateId": 78,
    }


@pytest.fixture
def sample_subrecipients_lambda_payload() -> SubrecipientLambdaPayload:
    return SubrecipientLambdaPayload(
        organization=OrganizationObj(
            id=99,
            preferences=PreferencesObj(current_reporting_period_id=123),
        ),
        user=UserObj(id=456, email="foo@example.com"),
        outputTemplateId=789,
    )


@pytest.fixture
def invalid_json_content() -> str:
    return '{"subrecipients": [{"id": 1, "name": "subrecipient 3" invalid json'


@pytest.fixture
def no_subrecipients_key_json_content() -> Dict[str, Any]:
    return {"other_key": []}


@pytest.fixture
def no_subrecipients_list_json_content() -> Dict[str, Any]:
    return {"subrecipients": "not_a_list"}


@pytest.fixture
def empty_subrecipients_list_json_content() -> Dict[str, Any]:
    return {"subrecipients": []}


@pytest.fixture
def subrecipients_no_uploads() -> dict[str, list[dict[str, Any]]]:
    return {"subrecipients": [{"id": 1, "Name": "Bob"}]}


@pytest.fixture
def empty_subrecipient_treasury_template() -> openpyxl.Workbook:
    return openpyxl.load_workbook(_SAMPLE_SUBRECIPIENT_TEMPLATE_EMPTY)


@pytest.fixture
def sample_subrecipient_uploads_with_dates() -> Dict[str, Any]:
    return {
        "subrecipientUploads": [
            {"id": 1, "updatedAt": "2023-07-01T12:00:00Z"},
            {"id": 2, "updatedAt": "2023-07-02T12:00:00Z"},
            {"id": 3, "updatedAt": "2023-06-30T12:00:00Z"},
        ]
    }

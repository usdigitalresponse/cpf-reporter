import json
from tempfile import NamedTemporaryFile
from typing import Any, Dict, List
from unittest.mock import ANY, MagicMock, Mock, call, patch

import pytest
from aws_lambda_typing.context import Context
from openpyxl import Workbook

from src.functions.subrecipient_treasury_report_gen import (
    FIRST_BLANK_ROW_NUM,
    WORKSHEET_NAME,
    SubrecipientLambdaPayload,
    get_most_recent_upload,
    handle,
    process_event,
    upload_workbook,
    write_subrecipients_to_workbook,
)
from src.lib.treasury_generation_common import OrganizationObj


class TestHandleNoEventOrNoContext:
    @patch("src.functions.subrecipient_treasury_report_gen.get_logger")
    def test_handle_no_event_provided(
        self, mock_get_logger: Mock, valid_aws_typing_context: Context
    ) -> None:
        mock_logger = MagicMock()
        mock_get_logger.return_value = mock_logger

        handle(None, valid_aws_typing_context)
        mock_logger.exception.assert_called_once_with("Missing event or context")

    @patch("src.functions.subrecipient_treasury_report_gen.get_logger")
    def test_handle_no_context_provided(self, mock_get_logger: Mock) -> None:
        mock_logger = MagicMock()
        mock_get_logger.return_value = mock_logger
        event: dict[str, str] = {}
        context = None

        handle(event, context)

        mock_logger.exception.assert_called_once_with("Missing event or context")


class TestHandleIncompleteEventInput:
    @patch("src.functions.subrecipient_treasury_report_gen.get_logger")
    def test_handle_no_organization_id(
        self, mock_get_logger: Mock, valid_aws_typing_context: Context
    ) -> None:
        mock_logger = MagicMock()
        mock_get_logger.return_value = mock_logger

        handle(
            {
                "organization": {
                    "preferences": {"current_reporting_period_id": 1},
                },
                "user": {"id": 1, "email": "a@b.c"},
                "outputTemplateId": 1,
            },
            valid_aws_typing_context,
        )

        mock_logger.exception.assert_called_once_with(
            "Exception parsing Subrecipient event payload"
        )

    @patch("src.functions.subrecipient_treasury_report_gen.get_logger")
    def test_handle_no_preferences(
        self, mock_get_logger: Mock, valid_aws_typing_context: Context
    ) -> None:
        mock_logger = MagicMock()
        mock_get_logger.return_value = mock_logger

        handle({"organization": {"id": 1}}, valid_aws_typing_context)

        mock_logger.exception.assert_called_once_with(
            "Exception parsing Subrecipient event payload"
        )

    @patch("src.functions.subrecipient_treasury_report_gen.get_logger")
    def test_handle_no_current_reporting_period_id(
        self, mock_get_logger: Mock, valid_aws_typing_context: Context
    ) -> None:
        mock_logger = MagicMock()
        mock_get_logger.return_value = mock_logger

        handle({"organization": {"id": 1, "preferences": {}}}, valid_aws_typing_context)

        mock_logger.exception.assert_called_once_with(
            "Exception parsing Subrecipient event payload"
        )

    @patch("src.functions.subrecipient_treasury_report_gen.get_logger")
    def test_handle_no_output_template_id(
        self, mock_get_logger: Mock, valid_aws_typing_context: Context
    ) -> None:
        mock_logger = MagicMock()
        mock_get_logger.return_value = mock_logger

        handle(
            {
                "organization": {
                    "id": 1,
                    "preferences": {"current_reporting_period_id": 1},
                }
            },
            valid_aws_typing_context,
        )

        mock_logger.exception.assert_called_once_with(
            "Exception parsing Subrecipient event payload"
        )


class TestInvalidSubrecipientsFile:
    @patch("src.functions.subrecipient_treasury_report_gen.boto3.client")
    @patch("src.functions.subrecipient_treasury_report_gen.tempfile.NamedTemporaryFile")
    def test_invalid_subrecipients_file(
        self,
        mock_tempfile: Mock,
        mock_boto_client: Mock,
        invalid_json_content: str,
        sample_subrecipients_lambda_payload: SubrecipientLambdaPayload,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        monkeypatch.setenv("REPORTING_DATA_BUCKET_NAME", "test-cpf-reporter")
        mock_s3_client = MagicMock()
        mock_boto_client.return_value = mock_s3_client
        mock_logger = MagicMock()

        temp_file = NamedTemporaryFile(delete=False)
        temp_file.write(invalid_json_content.encode())
        temp_file.seek(0)
        mock_tempfile.return_value.__enter__.return_value = temp_file

        process_event(sample_subrecipients_lambda_payload, mock_logger)

        mock_logger.exception.assert_called_once_with(
            "Subrecipients file for organization 99 and reporting period 123 does not contain valid JSON"
        )

    @patch("src.functions.subrecipient_treasury_report_gen.boto3.client")
    @patch("src.functions.subrecipient_treasury_report_gen.tempfile.NamedTemporaryFile")
    def test_no_subrecipients_key(
        self,
        mock_tempfile: Mock,
        mock_boto_client: Mock,
        no_subrecipients_key_json_content: str,
        sample_subrecipients_lambda_payload: SubrecipientLambdaPayload,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        monkeypatch.setenv("REPORTING_DATA_BUCKET_NAME", "test-cpf-reporter")
        mock_s3_client = MagicMock()
        mock_boto_client.return_value = mock_s3_client
        mock_logger = MagicMock()

        temp_file = NamedTemporaryFile(delete=False)
        temp_file.write(json.dumps(no_subrecipients_key_json_content).encode())
        temp_file.seek(0)
        mock_tempfile.return_value.__enter__.return_value = temp_file

        process_event(sample_subrecipients_lambda_payload, mock_logger)

        mock_logger.warning.assert_called_once_with(
            "Subrecipients file for organization 99 and reporting period 123 does not have any subrecipients listed"
        )

    @patch("src.functions.subrecipient_treasury_report_gen.boto3.client")
    @patch("src.functions.subrecipient_treasury_report_gen.tempfile.NamedTemporaryFile")
    def test_no_subrecipients_list(
        self,
        mock_tempfile: Mock,
        mock_boto_client: Mock,
        no_subrecipients_list_json_content: str,
        sample_subrecipients_lambda_payload: SubrecipientLambdaPayload,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        monkeypatch.setenv("REPORTING_DATA_BUCKET_NAME", "test-cpf-reporter")
        mock_s3_client = MagicMock()
        mock_boto_client.return_value = mock_s3_client
        mock_logger = MagicMock()

        temp_file = NamedTemporaryFile(delete=False)
        temp_file.write(json.dumps(no_subrecipients_list_json_content).encode())
        temp_file.seek(0)
        mock_tempfile.return_value.__enter__.return_value = temp_file

        process_event(sample_subrecipients_lambda_payload, mock_logger)

        mock_logger.warning.assert_called_once_with(
            "Subrecipients file for organization 99 and reporting period 123 does not have any subrecipients listed"
        )

    @patch("src.functions.subrecipient_treasury_report_gen.boto3.client")
    @patch("src.functions.subrecipient_treasury_report_gen.tempfile.NamedTemporaryFile")
    def test_empty_subrecipients_list(
        self,
        mock_tempfile: Mock,
        mock_boto_client: Mock,
        empty_subrecipients_list_json_content: str,
        sample_subrecipients_lambda_payload: SubrecipientLambdaPayload,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        monkeypatch.setenv("REPORTING_DATA_BUCKET_NAME", "test-cpf-reporter")
        mock_s3_client = MagicMock()
        mock_boto_client.return_value = mock_s3_client
        mock_logger = MagicMock()

        temp_file = NamedTemporaryFile(delete=False)
        temp_file.write(json.dumps(empty_subrecipients_list_json_content).encode())
        temp_file.seek(0)
        mock_tempfile.return_value.__enter__.return_value = temp_file

        process_event(sample_subrecipients_lambda_payload, mock_logger)

        mock_logger.warning.assert_called_once_with(
            "Subrecipients file for organization 99 and reporting period 123 does not have any subrecipients listed"
        )


class TestWriteSubrecipientsToWorkbook:
    def test_write_subrecipients_to_workbook_no_uploads(
        self,
        subrecipients_no_uploads: dict[str, list[dict[str, Any]]],
        empty_subrecipient_treasury_template: Workbook,
    ) -> None:
        mock_logger = MagicMock()

        write_subrecipients_to_workbook(
            subrecipients_no_uploads, empty_subrecipient_treasury_template, mock_logger
        )

        mock_logger.warning.assert_called_once_with(
            "Subrecipient in recent uploads file with id 1 and name Bob doesn't have any associated uploads, skipping in treasury report"
        )

    def test_write_subrecipients_to_workbook_empty_output_file_valid_subrecipients(
        self,
        valid_subrecipients_json_content: dict[str, list[dict[str, Any]]],
        empty_subrecipient_treasury_template: Workbook,
    ) -> None:
        mock_logger = MagicMock()

        write_subrecipients_to_workbook(
            valid_subrecipients_json_content,
            empty_subrecipient_treasury_template,
            mock_logger,
        )

        edited_row = empty_subrecipient_treasury_template["Baseline"][
            FIRST_BLANK_ROW_NUM
        ]
        subrecipient_upload_data = valid_subrecipients_json_content["subrecipients"][0][
            "subrecipientUploads"
        ][0]["rawSubrecipient"]

        assert edited_row[1].value == subrecipient_upload_data["Name"]
        assert (
            edited_row[2].value == subrecipient_upload_data["Recipient_Profile_ID__c"]
        )
        assert edited_row[3].value == subrecipient_upload_data["EIN__c"]
        assert (
            edited_row[4].value
            == subrecipient_upload_data["Unique_Entity_Identifier__c"]
        )
        assert edited_row[5].value == subrecipient_upload_data["POC_Name__c"]
        assert edited_row[6].value == subrecipient_upload_data["POC_Phone_Number__c"]
        assert edited_row[7].value == subrecipient_upload_data["POC_Email_Address__c"]
        assert edited_row[8].value == subrecipient_upload_data["Zip__c"]
        assert edited_row[9].value is None  # Zip4
        assert edited_row[10].value == subrecipient_upload_data["Address__c"]
        assert edited_row[11].value is None  # Address line 2
        assert edited_row[12].value is None  # Address line 3
        assert edited_row[13].value == subrecipient_upload_data["City__c"]
        assert edited_row[14].value == subrecipient_upload_data["State_Abbreviated__c"]


class TestUploadSubrecipientWorkbook:
    @patch("src.functions.subrecipient_treasury_report_gen.upload_generated_file_to_s3")
    @patch("src.functions.subrecipient_treasury_report_gen.convert_xlsx_to_csv")
    def test_upload_workbook(
        self,
        mock_convert_xlsx_to_csv: Mock,
        mock_upload_generated_file_to_s3: Mock,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        bucket_name = "test-cpf-reporter"
        monkeypatch.setenv("REPORTING_DATA_BUCKET_NAME", bucket_name)
        mock_s3_client = MagicMock()
        workbook = Workbook()
        workbook.create_sheet(WORKSHEET_NAME)
        organization_raw = {
            "id": 99,
            "preferences": {"current_reporting_period_id": 88},
        }
        organization = OrganizationObj.model_validate(organization_raw)

        upload_template_location_minus_filetype = (
            "treasuryreports/99/88/CPFSubrecipientTemplate"
        )
        upload_template_xlsx_key = f"{upload_template_location_minus_filetype}.xlsx"
        upload_template_csv_key = f"{upload_template_location_minus_filetype}.csv"

        upload_workbook(workbook, mock_s3_client, organization)

        mock_convert_xlsx_to_csv.assert_called_once()

        calls = [
            call(mock_s3_client, bucket_name, upload_template_xlsx_key, ANY),
            call(mock_s3_client, bucket_name, upload_template_csv_key, ANY),
        ]
        mock_upload_generated_file_to_s3.assert_has_calls(calls, any_order=True)


class TestGetMostRecentUpload:
    def test_get_most_recent_upload(
        self, sample_subrecipient_uploads_with_dates: dict[str, list[dict[str, Any]]]
    ) -> None:
        result = get_most_recent_upload(sample_subrecipient_uploads_with_dates)
        assert result["id"] == 2

    def test_get_most_recent_upload_with_single_entry(self) -> None:
        single_entry_subrecipient = {
            "subrecipientUploads": [{"id": 1, "updatedAt": "2023-07-01T12:00:00Z"}]
        }
        result = get_most_recent_upload(single_entry_subrecipient)
        assert result["id"] == 1

    def test_get_most_recent_upload_with_empty_list(self) -> None:
        empty_subrecipient: Dict[str, List[str]] = {"subrecipientUploads": []}
        with pytest.raises(IndexError):
            get_most_recent_upload(empty_subrecipient)

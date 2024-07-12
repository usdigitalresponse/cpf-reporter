from enum import Enum
from typing import IO, Any, Dict, Iterable, List, Optional, Tuple, Type, Union

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError
from pydantic_core import ErrorDetails

from src.lib.logging import get_logger
from src.schemas.project_types import ProjectType
from src.schemas.schema_versions import (
    CoverSheetRow,
    LogicSheetVersion,
    Project1ARow,
    Project1BRow,
    Project1CRow,
    SubrecipientRow,
    Version,
    getCoverSheetRowClass,
    getSchemaByProject,
    getSchemaMetadata,
    getSubrecipientRowClass,
    getVersionFromString,
)

type Errors = List[WorkbookError]
type Subrecipients = List[SubrecipientRow]

LOGIC_SHEET = "Logic"
COVER_SHEET = "Cover"
PROJECT_SHEET = "Project"
SUBRECIPIENTS_SHEET = "Subrecipients"
INITIAL_STARTING_ROW = 12

_logger = get_logger(__name__)


# enum for error level
# error, warning, info
# default is error
class ErrorLevel(Enum):
    ERR: int = 1
    WARN: int = 2
    INFO: int = 3


class WorkbookError:
    message: str
    row: str
    col: str
    tab: str
    field_name: str
    severity: str = ErrorLevel.WARN.name

    def __init__(
        self,
        message: str,
        row: str = "N/A",
        col: str = "N/A",
        tab: str = "N/A",
        field_name: str = "N/A",
        severity: str = ErrorLevel.WARN.name,
    ):
        self.message = message
        self.row = row
        self.col = col
        self.tab = tab
        self.field_name = field_name
        self.severity = severity


def map_values_to_headers(headers: Tuple, values: Iterable[Any]):
    return dict(zip(headers, values))


def is_empty_row(row_values: Tuple):
    return all(value in (None, "") for value in row_values)


def get_headers(sheet: Worksheet, cell_range: str) -> tuple:
    return tuple(header_cell.value for header_cell in sheet[cell_range][0])


def get_project_use_code(
    cover_sheet: Worksheet,
    version_string: str,
    row_dict: Optional[Dict[str, str]] = None,
) -> ProjectType:
    """This function gets the Project Use Code / Expenditure Category Group
    (depending on version) from the cover sheet or uses the row provided.
    If the project use code is not found or if it does not match any of the ProjectType
    values, it raises an error.
    """
    metadata = getSchemaMetadata(version_string)
    if not row_dict:
        cover_header = get_headers(cover_sheet, metadata["Cover"]["header_range"])
        cover_row = map(lambda cell: cell.value, cover_sheet[2])
        row_dict = map_values_to_headers(cover_header, cover_row)
    version = getVersionFromString(version_string)
    codeKey = "Expenditure Category Group"
    if version != Version.V2024_05_24:
        codeKey = "Project Use Code"
    code = row_dict[codeKey]
    return ProjectType.from_project_name(code)


def generate_error_text(
    field_name: str,
    error: ErrorDetails,
) -> str:
    """This function converts a list of ValidationError records for a single row
    into a list of WorkbookError records.

    Since each row has many different fields and there can be multiple errors in
    a single row, it maps the thrown error to the impacted column in the spreadsheet.
    It does so by first getting the error's location (the loc property), which is
    the field name, and then grabbing that field off of the relevant model class
    passed in.

    On the model class, the field definition has a property of `json_schema_extra`,
    defined in the schema on a per-field basis, which contains the column for that
    particular field.
    """
    error_type = error["type"]
    input = error["input"]
    if isinstance(input, str):
        input = input.strip()
    if error_type == "missing" or input in [None, ""]:
        if (
            field_name == "project_use_code"
            or field_name == "expenditure_category_group"
        ):
            return "EC code must be set"
        else:
            return f"Value is required for {field_name}"
    elif error_type == "string_too_long" or error_type == "string_too_short":
        return error["msg"].replace("String", field_name, 1)
    elif error_type in ["decimal_whole_digits", "decimal_max_places", "decimal_type"]:
        return error["msg"].replace("Decimal", field_name, 1)
    elif error_type in [
        "int_type",
        "string_type",
        "datetime_from_date_parsing",
        "decimal_parsing",
        "int_parsing",
    ]:
        return error["msg"].replace("Input", field_name, 1)
    else:
        return f'Error in field {field_name}-{error["msg"]}'


def get_workbook_errors_for_row(
    SheetModelClass: Type[
        Union[CoverSheetRow, Project1ARow, Project1BRow, Project1CRow, SubrecipientRow]
    ],
    e: ValidationError,
    row_num: int,
    sheet_name: str,
) -> List[WorkbookError]:
    workbook_errors: List[WorkbookError] = []
    for error in e.errors():
        """
            Sample structure of the variable `e` here:
            https://docs.pydantic.dev/latest/api/pydantic_core/#pydantic_core.ErrorDetails
            {
                'type': 'string_type',
                'loc': ('Identification_Number__c',),
                'msg': 'Input should be a valid string',
                'input': None,
                'url': 'https://errors.pydantic.dev/2.6/v/string_type'
            }
        """
        erroring_field_name: str = f'{error["loc"][0]}'
        try:
            field_instance = SheetModelClass.model_fields[erroring_field_name]
            if isinstance(field_instance.json_schema_extra, dict):
                erroring_column = field_instance.json_schema_extra["column"]
            else:
                raise Exception(
                    f"Issue with schema definition for field {erroring_field_name}."
                )
        except Exception as exception:
            _logger.error(
                f"Encountered unexpected exception while getting column for field {erroring_field_name} with error {error}. Details: {exception}"
            )
            erroring_column = "Unknown"

        message = generate_error_text(erroring_field_name, error)
        workbook_errors.append(
            WorkbookError(
                message=message,
                row=f"{row_num}",
                col=f"{erroring_column}",
                tab=sheet_name,
                field_name=erroring_field_name,
                severity=ErrorLevel.ERR.name,
            )
        )
    return workbook_errors


def validate(workbook: IO[bytes]) -> Tuple[Errors, Optional[str], Subrecipients]:
    """Validates a given Excel workbook according to CPF validation rules.

    Args:
        workbook: The Excel workbook file to read and validate.
    """

    """TemporaryFile
    1. Load workbook in read-only mode
    See: https://openpyxl.readthedocs.io/en/stable/optimized.html
    If there is any trouble loading files from memory please see here: https://stackoverflow.com/questions/20635778/using-openpyxl-to-read-file-from-memory
    """
    try:
        wb = load_workbook(filename=workbook, read_only=True)
        return validate_workbook(wb)
    except Exception as e:
        _logger.error(f"Unexpected Validation Error: {e}")
        return (
            [
                WorkbookError(
                    "Unable to validate workbook. Please reach out to grants-helpdesk@usdigitalresponse.org",
                    severity=ErrorLevel.ERR.name,
                )
            ],
            "Unkown",
        )

    finally:
        workbook.close()


def validate_workbook(workbook: Workbook) -> Tuple[Errors, Optional[str]]:
    """Validates a given Excel workbook according to CPF validation rules.

    Args:
        workbook: The Excel workbook to validate.
    """
    """
    0. Validate that the workbook has all the appropriate sheets
    """
    errors: Errors = []
    errors += validate_workbook_sheets(workbook)

    """
    1. Validate logic sheet to make sure the sheet has an appropriate version
    """
    logic_errors, version_string = validate_logic_sheet(workbook[LOGIC_SHEET])
    errors += logic_errors

    """
    2. Validate cover sheet and project selection. Pick the appropriate validator for the next step.
    """
    cover_errors, project_schema, project_use_code = validate_cover_sheet(
        workbook[COVER_SHEET], version_string
    )
    errors += cover_errors

    """
    3. Ensure all project rows are validated with the schema
    """
    if project_schema:
        project_errors, projects = validate_project_sheet(
            workbook[PROJECT_SHEET], project_schema, version_string
        )
        errors += project_errors
    else:
        projects = None

    """
    4. Ensure all subrecipient rows are validated with the schema
    """
    subrecipient_errors, subrecipients = validate_subrecipient_sheet(
        workbook[SUBRECIPIENTS_SHEET], version_string
    )
    errors += subrecipient_errors

    """
    5. Ensure all projects are mapped to a valid subrecipient
    """
    if projects:
        errors += validate_projects_subrecipients(
            projects, subrecipients, version_string
        )

    
    subrecipients = [subrecipient.model_dump() for subrecipient in subrecipients]

    return (errors, project_use_code, subrecipients)


def validate_workbook_sheets(workbook: Workbook) -> Errors:
    errors = []
    expected_sheets = [LOGIC_SHEET, COVER_SHEET, PROJECT_SHEET, SUBRECIPIENTS_SHEET]
    for sheet in expected_sheets:
        if sheet not in workbook.sheetnames:
            errors.append(
                WorkbookError(
                    f"Workbook is missing expected sheet: {sheet}",
                    severity=ErrorLevel.ERR.name,
                )
            )
    return errors


# Accepts a workbook from openpyxl
def validate_logic_sheet(logic_sheet: Worksheet) -> Tuple[Errors, str]:
    errors = []
    version_string = logic_sheet["B1"].value
    try:
        # Cell B1 contains the version
        LogicSheetVersion(**{"version": version_string})
    except ValidationError as e:
        for error in e.errors():
            errors.append(
                WorkbookError(
                    message=error["msg"].replace("Value error, ", ""),
                    row="1",
                    col="B",
                    tab=LOGIC_SHEET,
                    field_name="version",
                    severity=ErrorLevel.WARN.name,
                )
            )
    return errors, version_string


def validate_cover_sheet(
    cover_sheet: Worksheet, version_string: str
) -> Tuple[
    Errors,
    Optional[Type[Union[Project1ARow, Project1BRow, Project1CRow]]],
    Optional[ProjectType],
]:
    errors = []
    project_schema = None
    metadata = getSchemaMetadata(version_string)
    cover_header = get_headers(cover_sheet, metadata["Cover"]["header_range"])
    row_num = 2
    cover_row = map(lambda cell: cell.value, cover_sheet[row_num])
    row_dict = map_values_to_headers(cover_header, cover_row)
    CoverSheetRowClass = getCoverSheetRowClass(version_string)
    try:
        CoverSheetRowClass(**row_dict)
    except ValidationError as e:
        errors += get_workbook_errors_for_row(
            CoverSheetRowClass, e, row_num, COVER_SHEET
        )
        return (errors, None, None)

    try:
        project_use_code: ProjectType = get_project_use_code(
            cover_sheet, version_string, row_dict
        )
    except Exception as e:
        _logger.error(f"Unrecognized project use code: {e}")
        errors.append(
            WorkbookError(
                f"{COVER_SHEET} Sheet: Project Use Code is not recognized",
                "2",
                "B",
                COVER_SHEET,
                "Project Use Code",
                severity=ErrorLevel.ERR.name,
            )
        )
        return (errors, None, None)

    project_schema = getSchemaByProject(version_string, project_use_code)
    return (errors, project_schema, project_use_code)


def validate_project_sheet(
    project_sheet: Worksheet,
    project_schema: Type[Union[Project1ARow, Project1BRow, Project1CRow]],
    version_string: str,
) -> Tuple[Errors, List[Union[Project1ARow, Project1BRow, Project1CRow]]]:
    errors = []
    metadata = getSchemaMetadata(version_string)
    projects = []
    project_headers = get_headers(project_sheet, metadata["Project"]["header_range"])
    current_row = INITIAL_STARTING_ROW
    sheet_has_data = False
    for project_row in project_sheet.iter_rows(
        min_row=13, min_col=3, max_col=156, values_only=True
    ):
        current_row += 1
        if is_empty_row(project_row):
            continue
        sheet_has_data = True
        row_dict = map_values_to_headers(project_headers, project_row)
        try:
            # Note: we will need to set row number here on the project in order to display a meaningful error message.
            row_dict["row_num"] = current_row

            projects.append(project_schema(**row_dict))
        except ValidationError as e:
            errors += get_workbook_errors_for_row(
                project_schema, e, current_row, PROJECT_SHEET
            )

    if not sheet_has_data:
        errors += [
            WorkbookError(
                message="Upload doesn’t include any project records.",
                row=INITIAL_STARTING_ROW + 1,
                col=0,
                tab=PROJECT_SHEET,
                field_name="",
                severity=ErrorLevel.ERR.name,
            )
        ]

    return (errors, projects)


def validate_subrecipient_sheet(
    subrecipient_sheet: Worksheet, version_string: str
) -> Tuple[Errors, List[SubrecipientRow]]:
    errors = []
    metadata = getSchemaMetadata(version_string)
    subrecipients = []
    subrecipient_headers = get_headers(
        subrecipient_sheet, metadata["Subrecipients"]["header_range"]
    )
    current_row = INITIAL_STARTING_ROW
    SubrecipientRowClass = getSubrecipientRowClass(version_string)
    for subrecipient_row in subrecipient_sheet.iter_rows(
        min_row=13, min_col=3, max_col=16, values_only=True
    ):
        current_row += 1
        if is_empty_row(subrecipient_row):
            continue
        row_dict = map_values_to_headers(subrecipient_headers, subrecipient_row)
        try:
            subrecipients.append(SubrecipientRowClass(**row_dict))
        except ValidationError as e:
            errors += get_workbook_errors_for_row(
                SubrecipientRowClass, e, current_row, SUBRECIPIENTS_SHEET
            )
    return (errors, subrecipients)


def validate_projects_subrecipients(
    projects: List[Union[Project1ARow, Project1BRow, Project1CRow]],
    subrecipients: List[SubrecipientRow],
    version_string: str,
) -> Errors:
    # Versions older than V2024_05_24 don't have subrecipient information
    # in the project sheet so skip this validation.
    if getVersionFromString(version_string) != Version.active_version():
        return []

    errors = []
    subrecipients_by_uei_tin = {}
    for subrecipient in subrecipients:
        subrecipients_by_uei_tin[
            (subrecipient.EIN__c, subrecipient.Unique_Entity_Identifier__c)
        ] = subrecipient

    for project in projects:
        if (
            subrecipients_by_uei_tin.get(
                (project.Subrecipient_TIN__c, project.Subrecipient_UEI__c)
            )
            is None
        ):
            col_name_tin = project.__class__.model_fields[
                "Subrecipient_TIN__c"
            ].json_schema_extra["column"]
            col_name_uei = project.__class__.model_fields[
                "Subrecipient_UEI__c"
            ].json_schema_extra["column"]
            errors.append(
                WorkbookError(
                    message="You must submit a subrecipient record with the same UEI & TIN numbers entered for this project",
                    row=project.row_num,
                    col=f"{col_name_uei}, {col_name_tin}",
                    tab="Project",
                    field_name="Subrecipient_TIN__c and Subrecipient_UEI__c",
                    severity=ErrorLevel.ERR.name,
                )
            )
    return errors


if __name__ == "__main__":
    """Main block to test validate function

    Run with `poetry run python -m src.lib.workbook_validator upload.xlsm`
    """
    import sys

    file_path = sys.argv[1]
    with open(file_path, "rb") as f:
        errors, project_use_code = validate(f)
        if errors:
            print("Errors found:")
            for error in errors:
                print(error.message)
        else:
            print("No errors found")

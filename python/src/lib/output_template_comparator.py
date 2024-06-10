# Basic structure of output_template_comparator.py

# Returns a list of messages that contain differences between latest set of files and previous set of files.
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Set

from openpyxl import load_workbook

HEADER_ROW_INDEX = 4


@dataclass
class CPFDiffReport:
    new_sheets: Dict[str, List[str]] = field(default_factory=dict)
    removed_sheets: Dict[str, List[str]] = field(default_factory=dict)
    row_count_changed: Dict[str, List[str]] = field(default_factory=dict)
    column_count_changed: Dict[str, List[str]] = field(default_factory=dict)
    column_differences: Dict[str, List[str]] = field(default_factory=dict)
    cell_value_changed: Dict[str, List[str]] = field(default_factory=dict)
    new_files: List[str] = field(default_factory=list)
    removed_files: List[str] = field(default_factory=list)

    def __post_init__(self):
        self.new_sheets = defaultdict(list)
        self.removed_sheets = defaultdict(list)
        self.row_count_changed = defaultdict(list)
        self.column_count_changed = defaultdict(list)
        self.column_differences = defaultdict(list)
        self.cell_value_changed = defaultdict(list)

    def summary_report(self):
        return f"""
        New files: {self._format_item(self.new_files)}
        Removed files: {self._format_item(self.removed_files)}
        New sheets: {self._format_item(self.new_sheets)}
        Removed sheets: {self._format_item(self.removed_sheets)}
        Row count changed: {self._format_item(self.row_count_changed)}
        Column count changed: {self._format_item(self.column_count_changed)}
        Column differences: {self._format_item(self.column_differences)}
        Cell value changed: {self._format_item(self.cell_value_changed)}
        """

    @staticmethod
    def _format_item(item: any):
        if isinstance(item, list):
            return "\n".join(item)
        elif isinstance(item, dict):
            return "\n".join([f"{k}: {v}" for k, v in item.items()])
        else:
            return item


class CPFFileArchive:
    """Class for working with a CPF file archive"""

    _zip_file: zipfile.ZipFile
    _file_name_map: dict

    def __init__(self, zip_file: zipfile.ZipFile):
        self._zip_file = zip_file
        self._normalize_names()

    def normalized_file_names(self) -> Set[str]:
        """Returns a dictionary of normalized file names"""
        return set(self._file_name_map.keys())

    def file_by_name(self, name) -> zipfile.ZipExtFile:
        """Returns a file object by name"""
        return self._zip_file.open(self._file_name_map[name])

    def _normalize_names(self):
        """Normalizes file names"""
        suffix_regex = r"\s\(\d+\)"
        normalized_files = {}
        for file in self._zip_file.namelist():
            normalized_filename = re.sub(suffix_regex, "", file.split("/")[-1])
            normalized_files[normalized_filename] = file
        self._file_name_map = normalized_files


def compare_workbooks(
    latest_archive: CPFFileArchive, previous_archive: CPFFileArchive
) -> tuple:
    latest_files = latest_archive.normalized_file_names()
    previous_files = previous_archive.normalized_file_names()
    new_files = latest_files - previous_files
    removed_files = previous_files - latest_files
    common_files = latest_files & previous_files
    return common_files, new_files, removed_files


def compare_sheet_columns(previous_sheet, latest_sheet):
    previous_column_headers = [
        previous_sheet.cell(row=HEADER_ROW_INDEX, column=column).value
        for column in range(1, previous_sheet.max_column + 1)
    ]
    latest_column_headers = [
        latest_sheet.cell(row=HEADER_ROW_INDEX, column=column).value
        for column in range(1, latest_sheet.max_column + 1)
    ]

    added_columns = set(latest_column_headers) - set(previous_column_headers)
    removed_columns = set(previous_column_headers) - set(latest_column_headers)
    columns_in_both = set(latest_column_headers) & set(previous_column_headers)
    # build a dict of key column header, value two-tuple of previous and latest column values
    header_map = {
        header: (
            previous_column_headers.index(header) + 1,
            latest_column_headers.index(header) + 1,
        )
        for header in columns_in_both
    }
    return added_columns, removed_columns, header_map


def compare_cell_values(previous_sheet, latest_sheet, header_map):
    differences = []
    for row in range(HEADER_ROW_INDEX, latest_sheet.max_row + 1):
        for header, (previous_column, latest_column) in header_map.items():
            previous_value = previous_sheet.cell(row=row, column=previous_column).value
            latest_value = latest_sheet.cell(row=row, column=latest_column).value
            if previous_value != latest_value:
                differences.append(
                    f"{header} - Row {row} - Previous: {previous_value}, Latest: {latest_value}"
                )
    return differences


def compare(
    latest_zip_files: zipfile.ZipFile, previous_zip_files: zipfile.ZipFile
) -> List[str]:
    # read the files from the zip files
    # compare the files
    # return the differences
    differences = CPFDiffReport()

    # first find any files that are in the latest set that are not in the previous set
    latest_archive = CPFFileArchive(latest_zip_files)
    previous_archive = CPFFileArchive(previous_zip_files)

    common_files, new_files, removed_files = compare_workbooks(
        latest_archive, previous_archive
    )
    differences.new_files += list(new_files)
    differences.removed_files += list(removed_files)

    for file in common_files:
        latest_file = latest_archive.file_by_name(file)
        previous_file = previous_archive.file_by_name(file)

        latest_workbook = load_workbook(latest_file)
        previous_workbook = load_workbook(previous_file)

        latest_sheets = latest_workbook.sheetnames
        previous_sheets = previous_workbook.sheetnames

        new_sheets = set(latest_sheets) - set(previous_sheets)
        for sheet in new_sheets:
            differences.new_sheets[file].append(sheet)

        removed_sheets = set(previous_sheets) - set(latest_sheets)
        for sheet in removed_sheets:
            differences.removed_sheets[file].append(sheet)

        common_sheets = set(latest_sheets) & set(previous_sheets)
        for sheet in common_sheets:
            latest_sheet = latest_workbook[sheet]
            previous_sheet = previous_workbook[sheet]

            if latest_sheet.max_row != previous_sheet.max_row:
                differences.row_count_changed[f"{file}: {sheet}"] = (
                    f"{latest_sheet.max_row} rows in latest, {previous_sheet.max_row} rows in previous"
                )

            if latest_sheet.max_column != previous_sheet.max_column:
                differences.column_count_changed[f"{file}: {sheet}"] = (
                    f"{latest_sheet.max_column} columns in latest, {previous_sheet.max_column} columns in previous"
                )

            added_columns, removed_columns, header_map = compare_sheet_columns(
                previous_sheet, latest_sheet
            )
            if added_columns:
                differences.column_differences[f"{file}: {sheet}"] += [
                    f"Added columns: {', '.join(added_columns)}"
                ]
            if removed_columns:
                differences.column_differences[f"{file}: {sheet}"] += [
                    f"Removed columns: {', '.join(removed_columns)}"
                ]

            cell_value_differences = compare_cell_values(
                previous_sheet, latest_sheet, header_map
            )
            if cell_value_differences:
                differences.cell_value_changed[f"{file}: {sheet}"] += (
                    cell_value_differences
                )

    return differences


def load_files(zip_path) -> zipfile.ZipFile:
    """Loads XLSX files from a zip file.

    Args:
        zip_path (str): Path to the zip file

    Returns:
        list: List of files in the zip file
    """

    return zipfile.ZipFile(zip_path, "r")


if __name__ == "__main__":
    """Main function to test comparator function.

    Accepts two arguments:
        1. Argument is the "latest" template. Second argument is the previous template.
        2. Returns a list of changes that are in the latest template that are different
        from the previous template.

    Run with `poetry run python -m src.lib.output_template_comparator 2024_05_16.zip 2024_04_26.zip`
    """
    import sys

    latest_zip_path = sys.argv[1]
    previous_zip_path = sys.argv[2]

    latest_zip_file = load_files(latest_zip_path)
    previous_zip_file = load_files(previous_zip_path)
    differences = compare(latest_zip_file, previous_zip_file)

    print(differences.summary_report())

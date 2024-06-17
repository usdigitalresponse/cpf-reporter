from datetime import datetime
from openpyxl import load_workbook
from typing import IO, List, Union, Dict
from urllib.parse import unquote_plus
import csv
import json
import re
import tempfile
from typing import Optional

import boto3
import structlog
from aws_lambda_typing.context import Context
from aws_lambda_typing.events import S3Event
from mypy_boto3_s3.client import S3Client

from src.lib.logging import get_logger, reset_contextvars
from src.schemas.schema_versions import Version, getSchemaByProject, Project1ARow, Project1BRow, Project1CRow
from src.schemas.project_types import ProjectType

OUTPUT_STARTING_ROW = 8
PROJECT_USE_CODE = ProjectType._1C
VERSION = Version.V2024_05_24
ProjectRowSchema = getSchemaByProject(VERSION, PROJECT_USE_CODE)


@reset_contextvars
def handle(event: S3Event, context: Context):
    """Lambda handler for generating Treasure Reports for Project Code 1C

    Args:
        event: S3 Lambda event of type `s3:ObjectCreated:*`
        context: Lambda context
    """
    structlog.contextvars.bind_contextvars(
        lambda_event={"step_function": event}
    )
    logger = get_logger()
    logger.info("received new invocation event from step function")

    uploadsByExpenditureCategory = event.get("uploadsByExpenditureCategory", {})
    file_info_list = uploadsByExpenditureCategory.get(PROJECT_USE_CODE, {})

    project_id_to_data: Dict[str, List[Union[Project1ARow, Project1BRow, Project1CRow]]] = {}
    project_id_to_upload_date: Dict[str, datetime] = {}
    for _, file_info in file_info_list.items():
        # Download projects from DynamoDB
        projects, createdAt = get_project_from_db(

        )
        combine_project_rows(
            projects = projects,
            project_id_to_upload_date = project_id_to_upload_date,
            project_id_to_data = project_id_to_data,
            createdAt = createdAt,
        )

    with tempfile.NamedTemporaryFile() as file:
        s3_client: S3Client = boto3.client("s3")
        # Get the output template xlsx
        download_workbook(
            s3_client,
            event["Records"][0]["s3"]["bucket"]["name"],
            event["Records"][0]["s3"]["object"]["key"],
            file,
        )
        workbook = load_workbook(filename=file)
        populate_output_report(workbook=workbook, project_id_to_data=project_id_to_data)

        with tempfile.NamedTemporaryFile("w") as csv_file:
            convert_xlsx_to_csv(csv_file, workbook)
            upload_generated_file_to_s3(
                s3_client,
                event["Records"][0]["s3"]["bucket"]["name"],
                f"{event["Records"][0]["s3"]["object"]["key"]}.json",
                csv_file,
            )


def get_project_from_db():
    return [], datetime.now()

def combine_project_rows(
        projects: List[Union[Project1ARow, Project1BRow, Project1CRow]],
        project_id_to_upload_date: Dict[str, datetime],
        project_id_to_data: Dict[str, List[Union[Project1ARow, Project1BRow, Project1CRow]]],
        createdAt: datetime,
):
    """
    Combine projects together and check for conflicts.
    If there is a conflict, choose the most recent project based on createdAt time.
    """
    # Get project rows from workbook
    for project_json in projects:
        project = ProjectRowSchema(**project_json)
        # Make sure to attach the date of the `UploadObject.createdAt` to the row somehow
        # to choose the most recent project record in case of conflicts
        existing_date = project_id_to_upload_date.get(project.Identification_Number__c)
        if existing_date and existing_date > createdAt:
            continue
        else:
            project_id_to_upload_date[project.Identification_Number__c] = createdAt
            project_id_to_data[project.Identification_Number__c] = project


def populate_output_report(
        workbook: IO[bytes],
        project_id_to_data: Dict[str, List[Union[Project1ARow, Project1BRow, Project1CRow]]]
):
    """
    Append projects to the xlsx file
    """
    # Append the project rows
    sheet = workbook['Baseline']

    row_num = OUTPUT_STARTING_ROW
    for _, row in project_id_to_data.items():
        row_schema = row.model_json_schema()['properties']
        row_dict = dict(row)

        row_with_output_cols = {}
        for prop in row_dict.keys():
            prop_meta = row_schema.get(prop)
            if not prop_meta:
                raise Exception('Property not found. Cannot generate report')
            if prop_meta[f'output_column_{PROJECT_USE_CODE}']:
                row_with_output_cols[prop_meta[f'output_column_{PROJECT_USE_CODE}']] = row_dict[prop]

        for col in row_with_output_cols.keys():
            sheet[f'{col}{row_num}'] = str(row_with_output_cols[col])

        row_num += 1

    return row_num


def escape_for_csv(text: Optional[str]):
    if not text:
        return text
    text = text.replace("\n", " -- ")
    text = text.replace("\r", " -- ")
    return text


def convert_xlsx_to_csv(csv_file: IO[bytes], file: IO[bytes], num_rows):
    """
    Convert xlsx file to csv.
    """
    sheet = file['Baseline']
    csv_file_handle = csv.writer(csv_file, delimiter=",")

    row_num = 1
    for eachrow in sheet.rows:
        if row_num >= num_rows:
            break
        csv_file_handle.writerow([escape_for_csv(cell.value) for cell in eachrow])
        row_num = row_num + 1
    
    return csv_file


def download_workbook(client: S3Client, bucket: str, key: str, destination: IO[bytes]):
    """Downloads an S3 object to a local file.

    Args:
        client: Client facilitating download from S3
        bucket: Name of the S3 bucket containing the workbook
        key: S3 object key for the file to download
        destination: File-like object (in binary mode) where the S3 object will be written
    """
    logger = get_logger()
    logger.debug("downloading workbook from s3")

    try:
        client.download_fileobj(bucket, unquote_plus(key), destination)
    except:
        logger.exception("failed to download workbook from S3")
        raise


def upload_generated_file_to_s3(client: S3Client, bucket, key: str, file: IO[bytes]):
    """Persists workbook validation results to S3.

    Args:
        client: Client facilitating upload to S3
        results: Errors resulting from workbook validation.
    """
    logger = get_logger()

    logger = logger.bind(
        upload={"s3": {"bucket": bucket, "key": key}}
    )
    try:
        client.put_object(
            file,
            Bucket=bucket,
            Key=unquote_plus(key),
            ServerSideEncryption="AES256",
        )
    except:
        logger.exception("failed to upload treasury report to S3")
        raise

    logger.info("successfully uploaded treasury report results to s3")

import json
import os
import tempfile
from datetime import datetime
from typing import IO, Dict, List, Set, Union

import boto3
import structlog
from aws_lambda_typing.context import Context
from botocore.exceptions import ClientError
from mypy_boto3_s3.client import S3Client
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from src.lib.logging import get_logger, reset_contextvars
from src.lib.s3_helper import download_s3_object, upload_generated_file_to_s3
from src.lib.treasury_generation_common import (
    OrganizationObj,
    OutputFileType,
    UserObj,
    get_generated_output_file_key,
    get_output_template,
)
from src.lib.workbook_utils import convert_xlsx_to_csv
from src.lib.workbook_validator import validate_project_sheet
from src.schemas.project_types import ProjectType
from src.schemas.schema_versions import (
    Project1ARow,
    Project1BRow,
    Project1CRow,
    Version,
    getSchemaByProject,
)

OUTPUT_STARTING_ROW = 8
VERSION = Version.active_version()
PROJECT_SHEET = "Project"


class UploadObj(BaseModel):
    objectKey: str
    filename: str
    createdAt: datetime


type AgencyId = str


class ProjectLambdaPayload(BaseModel):
    organization: OrganizationObj
    user: UserObj
    outputTemplateId: int
    ProjectType: str
    uploadsToAdd: Dict[AgencyId, UploadObj]
    uploadsToRemove: Dict[AgencyId, UploadObj]


@reset_contextvars
def handle(event: ProjectLambdaPayload, context: Context):
    """Lambda handler for generating Treasury Reports

    This function creates/outputs 3 files (all with the same name but different
    extensions):
    - XLSX file which contains the output for internal use
    - CSV version of the XLSX file for treasury use
    - JSON Metadata which maps the {project_id}_{agency_id} to the row in the
        XLSX file to track duplicate entries.

    Args:
        event: S3 Lambda event of type `s3:ObjectCreated:*`
        context: Lambda context
    """
    structlog.contextvars.bind_contextvars(lambda_event={"step_function": event})
    logger = get_logger()
    logger.info("received new invocation event from step function")

    if not event or not context:
        logger.exception("Missing event or context")
        return {"statusCode": 400, "body": "Bad Request - No event or context"}

    try:
        payload = ProjectLambdaPayload.model_validate(event)
    except Exception:
        logger.exception("Exception parsing Project event payload")
        return {"statusCode": 400, "body": "Bad Request"}

    try:
        process_event(payload, logger)
    except Exception:
        logger.exception("Exception processing Project file generation event")
        return {"statusCode": 500, "body": "Internal Server Error"}

    return {"statusCode": 200, "body": "Success"}


def process_event(payload: ProjectLambdaPayload, logger: structlog.stdlib.BoundLogger):
    """
    This function is structured as followed:
    1) Load the metadata
    2) Download the existing output XLSX file. If it doesn't exist, download
        the template file.
    3) Open files in the outdatedUploadsByExpenditureCategory parameter. These
        files contain projects that we want to remove from the output file.
        Then remove those files from the output file while simulateneously
        updating the metadata (when you delete a row in the XLSX, the metadata
        needs to be updated).
    4) Open files in the uploadsByExpenditureCategory parameter. These files
        contain projects that we want to add to the output file. There is some
        complicated logic to handle cases where there are duplicate project
        entries from the SAME agency. When that happens, we want to use the
        latest entry, denoted by the UploadObject's createdAt data. OR if there
        is already an entry in the output file, overwrite it since presumably
        the new worksheet has the updated data.
    5) Save the XLSX sheet, convert it to CSV and upload it, save metadata to
        s3.
    """
    s3_client: S3Client = boto3.client("s3")

    project_use_code = None
    try:
        project_use_code = ProjectType.from_project_name(payload.ProjectType)
    except ValueError:
        logger.error(
            f"Project name '{payload.ProjectType}' is not a recognized project type."
        )
        return {"statusCode": 400, "body": "Invalid project name"}

    ProjectRowSchema = getSchemaByProject(VERSION, project_use_code)

    organization = payload.organization

    # If the treasury report file exists, download it and store the data
    # If it doesn't exist, download the output template

    ### 1) Load the metadata
    project_agency_id_to_row_map = get_existing_output_metadata(
        s3_client=s3_client,
        organization=organization,
        project_use_code=project_use_code,
        logger=logger,
    )

    ### 2) Download the existing output XLSX file. If it doesn't exist, download
    #    the template file.
    with tempfile.NamedTemporaryFile() as output_file:
        highest_row_num = download_output_file(
            s3_client=s3_client,
            organization=organization,
            project_use_code=project_use_code,
            project_agency_id_to_row_map=project_agency_id_to_row_map,
            output_file=output_file,
            output_template_id=payload.outputTemplateId,
            logger=logger,
        )

        output_workbook = load_workbook(filename=output_file, read_only=False)
        output_sheet = output_workbook["Baseline"]

        ### 3) Open files in the uploadsToRemove parameter.
        # If outdated file info list is provided, grab the project id agency id to remove
        project_agency_ids_to_remove = get_outdated_projects_to_remove(
            s3_client=s3_client,
            uploads_by_agency_id=payload.uploadsToRemove,
            ProjectRowSchema=ProjectRowSchema,
            logger=logger,
        )

        # Delete rows corresponding to project_agency_ids_to_remove in the existing output file
        # Also update project_agency_id_to_row_map with the new row numbers
        highest_row_num = update_project_agency_ids_to_row_map(
            project_agency_id_to_row_map=project_agency_id_to_row_map,
            project_agency_ids_to_remove=project_agency_ids_to_remove,
            sheet=output_sheet,
            highest_row_num=highest_row_num,
        )

        ### 4) Open files in the uploadsByExpenditureCategory parameter.
        # Get project data to add to the output
        project_id_agency_id_to_upload_date: Dict[str, datetime] = {}
        for agency_id, file_info in payload.uploadsToAdd.items():
            with tempfile.NamedTemporaryFile() as file:
                # Download projects from S3
                created_at = file_info.createdAt

                # TODO: Add a verification to ensure that the objectKey here is formatted correctly.
                # Path must be: uploads/{organization_id}/{agency_id}/{reporting_period_id}/{upload_id}/{filename}.xlsm
                try:
                    download_s3_object(
                        s3_client,
                        os.environ["REPORTING_DATA_BUCKET_NAME"],
                        file_info.objectKey,
                        file,
                    )
                except ClientError as e:
                    error = e.response.get("Error") or {}
                    if error.get("Code") == "404":
                        logger.exception(
                            f"Expected to find an upload with key: {file_info.objectKey}"
                        )
                    raise
                # Load workbook
                wb = load_workbook(filename=file, read_only=True)
                # Get and combine project rows
                highest_row_num = combine_project_rows(
                    project_workbook=wb,
                    output_sheet=output_sheet,
                    project_use_code=project_use_code,
                    highest_row_num=highest_row_num,
                    ProjectRowSchema=ProjectRowSchema,
                    project_id_agency_id_to_upload_date=project_id_agency_id_to_upload_date,
                    project_id_agency_id_to_row_num=project_agency_id_to_row_map,
                    created_at=created_at,
                    agency_id=agency_id,
                )

        ### 5) Save data
        # Output XLSX file
        with tempfile.NamedTemporaryFile("r+") as new_output_file:
            output_workbook.save(new_output_file.name)
            upload_generated_file_to_s3(
                client=s3_client,
                bucket=os.environ["REPORTING_DATA_BUCKET_NAME"],
                key=get_generated_output_file_key(
                    file_type=OutputFileType.XLSX,
                    project=project_use_code,
                    organization=organization,
                ),
                file=new_output_file,
            )
        # Output CSV file for treasury
        with tempfile.NamedTemporaryFile("r+") as csv_file:
            convert_xlsx_to_csv(csv_file, output_workbook, highest_row_num)
            upload_generated_file_to_s3(
                client=s3_client,
                bucket=os.environ["REPORTING_DATA_BUCKET_NAME"],
                key=get_generated_output_file_key(
                    file_type=OutputFileType.CSV,
                    project=project_use_code,
                    organization=organization,
                ),
                file=csv_file,
            )
        # Store project_id_agency_id to row number in a json file
        with tempfile.NamedTemporaryFile("r+") as json_file:
            json.dump(project_agency_id_to_row_map, fp=json_file)
            upload_generated_file_to_s3(
                client=s3_client,
                bucket=os.environ["REPORTING_DATA_BUCKET_NAME"],
                key=get_generated_output_file_key(
                    file_type=OutputFileType.JSON,
                    project=project_use_code,
                    organization=organization,
                ),
                file=json_file,
            )

def download_output_file(
    s3_client: S3Client,
    output_file: IO[bytes],
    project_agency_id_to_row_map: Dict[str, int],
    organization: OrganizationObj,
    project_use_code: str,
    output_template_id: int,
    logger: structlog.stdlib.BoundLogger,
) -> int:
    highest_row_num = None
    if project_agency_id_to_row_map:
        """
        Output file already exists, download it
        """
        try:
            download_s3_object(
                client=s3_client,
                bucket=os.environ["REPORTING_DATA_BUCKET_NAME"],
                key=get_generated_output_file_key(
                    file_type=OutputFileType.XLSX,
                    project=project_use_code,
                    organization=organization,
                ),
                destination=output_file,
            )
        except ClientError as e:
            error = e.response.get("Error") or {}
            if error.get("Code") == "404":
                logger.exception("Expected to find an existing treasury output report")
                raise
        highest_row_num = max(project_agency_id_to_row_map.values())
    else:
        """
        Output file doesn't exist, download the empty template
        """
        get_output_template(
            s3_client=s3_client,
            output_template_id=output_template_id,
            project=project_use_code,
            destination=output_file,
        )
        highest_row_num = OUTPUT_STARTING_ROW - 1
    return highest_row_num

def get_existing_output_metadata(
    s3_client,
    organization: OrganizationObj,
    project_use_code: str,
    logger: structlog.stdlib.BoundLogger,
) -> Dict[str, int]:
    existing_project_agency_id_to_row_number = {}
    with tempfile.NamedTemporaryFile() as existing_file:
        try:
            download_s3_object(
                s3_client,
                os.environ["REPORTING_DATA_BUCKET_NAME"],
                get_generated_output_file_key(
                    file_type=OutputFileType.JSON,
                    project=project_use_code,
                    organization=organization,
                ),
                existing_file,
            )
        except ClientError as e:
            error = e.response.get("Error") or {}
            if error.get("Code") == "404":
                logger.info(
                    "There is no existing metadata file for this treasury report"
                )
                existing_file = None
            else:
                raise

        if existing_file:
            existing_project_agency_id_to_row_number = json.load(existing_file)

    return existing_project_agency_id_to_row_number


def get_projects_to_remove(
    workbook: Workbook,
    ProjectRowSchema: Union[Project1ARow, Project1BRow, Project1CRow],
    agency_id: str,
):
    """
    Get the set of '{project_id}_{agency_id}' ids to remove from the
    existing output file.
    """
    project_agency_ids_to_remove = set()
    # Get projects
    _, projects = validate_project_sheet(
        workbook[PROJECT_SHEET], ProjectRowSchema, VERSION
    )
    # Store projects to remove
    for project in projects:
        project_agency_ids_to_remove.add(
            f"{project.Identification_Number__c}_{agency_id}"
        )
    return project_agency_ids_to_remove


def get_outdated_projects_to_remove(
    s3_client,
    uploads_by_agency_id: Dict[AgencyId, UploadObj],
    ProjectRowSchema: Union[Project1ARow, Project1BRow, Project1CRow],
    logger: structlog.stdlib.BoundLogger,
):
    """
    Open the files in the outdated_file_info_list and get the projects to
    remove.
    """
    project_agency_ids_to_remove = set()
    for agency_id, file_info in uploads_by_agency_id.items():
        with tempfile.NamedTemporaryFile() as file:
            # Download projects from S3
            try:
                download_s3_object(
                    client=s3_client,
                    bucket=os.environ["REPORTING_DATA_BUCKET_NAME"],
                    key=file_info.objectKey,
                    destination=file,
                )
            except ClientError as e:
                error = e.response.get("Error") or {}
                if error.get("Code") == "404":
                    logger.exception(
                        f"Expected to find an upload with key: {file_info.objectKey}"
                    )
                raise
            # Load workbook
            outdated_wb = load_workbook(filename=file, read_only=True)
            project_agency_ids_to_remove = project_agency_ids_to_remove.union(
                get_projects_to_remove(
                    workbook=outdated_wb,
                    ProjectRowSchema=ProjectRowSchema,
                    agency_id=agency_id,
                )
            )
    return project_agency_ids_to_remove


def update_project_agency_ids_to_row_map(
    project_agency_id_to_row_map: Dict[str, int],
    project_agency_ids_to_remove: Set[str],
    highest_row_num: int,
    sheet: Worksheet,
):
    """
    Delete rows corresponding to project_agency_ids_to_remove in the existing
    output file. Also update project_agency_id_to_row_map with the new row
    numbers.
    """
    if project_agency_ids_to_remove:
        rows_to_remove = []
        for project_agency_id in project_agency_ids_to_remove:
            row_num = project_agency_id_to_row_map.get(project_agency_id, None)
            if row_num:
                rows_to_remove.append(row_num)
                project_agency_id_to_row_map.pop(project_agency_id)
        rows_to_remove.sort(reverse=True)

        for row_num in rows_to_remove:
            if sheet:
                sheet.delete_rows(row_num)
            highest_row_num = highest_row_num - 1
            for k, v in project_agency_id_to_row_map.items():
                if v > row_num:
                    project_agency_id_to_row_map[k] = v - 1
    return highest_row_num


def insert_project_row(
    project_use_code: str,
    sheet: Worksheet,
    row_num: int,
    row: Union[Project1ARow, Project1BRow, Project1CRow],
):
    """
    Append project to the xlsx file

    sheet is Optional only for tests
    """
    row_schema = row.model_json_schema()["properties"]
    row_dict = dict(row)

    row_with_output_cols = {}
    for prop in row_dict.keys():
        prop_meta = row_schema.get(prop)
        if not prop_meta:
            raise Exception("Property not found. Cannot generate report")
        if prop_meta[f"treasury_report_col_{project_use_code}"]:
            row_with_output_cols[
                prop_meta[f"treasury_report_col_{project_use_code}"]
            ] = row_dict[prop]

    for col in row_with_output_cols.keys():
        if sheet:
            sheet[f"{col}{row_num}"] = str(row_with_output_cols[col])


def combine_project_rows(
    project_workbook: Workbook,
    output_sheet: Worksheet,
    project_use_code: str,
    highest_row_num: int,
    ProjectRowSchema: Union[Project1ARow, Project1BRow, Project1CRow],
    project_id_agency_id_to_upload_date: Dict[str, datetime],
    project_id_agency_id_to_row_num: Dict[str, int],
    created_at: datetime,
    agency_id: str,
):
    """
    Combine projects together and check for conflicts.
    If there is a conflict, choose the most recent project based on created_at time.
    """
    # Get projects
    result = validate_project_sheet(
        project_workbook[PROJECT_SHEET], ProjectRowSchema, VERSION
    )
    projects: List[Union[Project1ARow, Project1BRow, Project1CRow]] = result[1]
    # Get project rows from workbook
    for project in projects:
        # Make sure to attach the date of the `UploadObject.createdAt` to the row somehow
        # to choose the most recent project record in case of conflicts
        project_agency_id = f"{project.Identification_Number__c}_{agency_id}"
        existing_date = project_id_agency_id_to_upload_date.get(project_agency_id)
        if existing_date and existing_date > created_at:
            continue
        else:
            project_id_agency_id_to_upload_date[project_agency_id] = created_at
            existing_project_row = project_id_agency_id_to_row_num.get(
                project_agency_id
            )
            if existing_project_row:
                row_to_insert = existing_project_row
            else:
                highest_row_num = highest_row_num + 1
                project_id_agency_id_to_row_num[project_agency_id] = highest_row_num
                row_to_insert = highest_row_num
            insert_project_row(
                project_use_code=project_use_code,
                sheet=output_sheet,
                row_num=row_to_insert,
                row=project,
            )
    return highest_row_num

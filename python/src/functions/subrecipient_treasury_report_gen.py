import json
import os
import tempfile
from datetime import datetime
from typing import Any, Dict

import boto3
import structlog
from aws_lambda_typing.context import Context
from botocore.exceptions import ClientError
from mypy_boto3_s3.client import S3Client
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from src.lib.logging import get_logger, reset_contextvars
from src.lib.s3_helper import download_s3_object, upload_generated_file_to_s3
from src.lib.treasury_generation_common import (
    OrganizationObj,
    OutputFileType,
    UserObj,
    get_generated_output_file_key,
    get_output_template,
)
from src.lib.workbook_utils import convert_xlsx_to_csv, find_last_populated_row
from src.schemas.schema_versions import getSubrecipientRowClass

FIRST_BLANK_ROW_NUM = 8
WORKSHEET_NAME = "Baseline"
FIRST_DATA_COLUMN = "B"


class SubrecipientLambdaPayload(BaseModel):
    organization: OrganizationObj
    user: UserObj
    outputTemplateId: int


@reset_contextvars
def handle(event: Dict[str, Any], context: Context):
    """Lambda handler for generating subrecipients file for treasury report

    Args:
        event: Step function that passes the following parameters:
        {
            organization: <all fields in organization object>,
            user: <id and email fields of the user object>,
            outputTemplateId: <id for the output template to use>
        }
        context: Lambda context

    """
    structlog.contextvars.bind_contextvars(
        lambda_event={"subrecipient_step_function": event}
    )
    logger = get_logger()
    logger.info("received new invocation event from step function")
    if not event or not context:
        logger.exception("Missing event or context")
        return {"statusCode": 400, "body": "Bad Request - No event or context"}

    try:
        payload = SubrecipientLambdaPayload.model_validate(event)
    except Exception:
        logger.exception("Exception parsing Subrecipient event payload")
        return {"statusCode": 400, "body": "Bad Request - payload validation failed"}

    try:
        process_event(payload, logger)
    except Exception:
        logger.exception("Exception processing Subrecipient file generation event")
        return {"statusCode": 500, "body": "Internal Server Error"}

    return {"statusCode": 200, "body": "Success"}


def process_event(
    payload: SubrecipientLambdaPayload, logger: structlog.stdlib.BoundLogger
):
    """
    This function should:
    1. Parse necessary inputs from the event
    2. Download a file of recent subrecipients from S3 to put into the output template
    3. Download the output template itself from S3
    4. Iterate through recent subrecipients and put their information into the output template
    5. Upload the output template, in both xlsx and csv formats, to S3
    """
    organization_id = payload.organization.id
    reporting_period_id = payload.organization.preferences.current_reporting_period_id
    output_template_id = payload.outputTemplateId

    s3_client: S3Client = boto3.client("s3")

    with tempfile.NamedTemporaryFile() as recent_subrecipients_file:
        with structlog.contextvars.bound_contextvars(
            subrecipients_filename=recent_subrecipients_file.name
        ):
            try:
                download_s3_object(
                    client=s3_client,
                    bucket=os.environ["REPORTING_DATA_BUCKET_NAME"],
                    key=f"treasuryreports/{organization_id}/{reporting_period_id}/subrecipients",
                    destination=recent_subrecipients_file,
                )
            except ClientError as e:
                error = e.response.get("Error") or {}
                if error.get("Code") == "404":
                    logger.info(
                        f"No subrecipients for organization {organization_id} and reporting period {reporting_period_id}"
                    )
                    return
                else:
                    raise

        recent_subrecipients_file.seek(0)

    recent_subrecipients = ...
    try:
        recent_subrecipients = json.load(recent_subrecipients_file)
    except json.JSONDecodeError:
        logger.exception(
            f"Subrecipients file for organization {organization_id} and reporting period {reporting_period_id} does not contain valid JSON"
        )
        return

    if no_subrecipients_in_file(recent_subrecipients=recent_subrecipients):
        logger.warning(
            f"Subrecipients file for organization {organization_id} and reporting period {reporting_period_id} does not have any subrecipients listed"
        )
        return

    output_file = tempfile.NamedTemporaryFile()
    get_output_template(s3_client, output_template_id, "Subrecipient", output_file)

    workbook = load_workbook(filename=output_file)
    write_subrecipients_to_workbook(
        recent_subrecipients=recent_subrecipients,
        workbook=workbook,
        logger=logger,
    )

    upload_workbook(workbook, s3_client, payload.organization)

    output_file.close()
    recent_subrecipients_file.close()

    return {
        "statusCode": 200,
        "Payload": {
            "organization_id": organization_id,
            "reporting_period_id": reporting_period_id,
        }
    }


def no_subrecipients_in_file(recent_subrecipients):
    """
    Helper method to determine if the recent_subrecipients JSON object in
    the recent subrecipients file downloaded from S3 has actual subrecipients in it or not
    """
    return (
        "subrecipients" not in recent_subrecipients
        or not isinstance(recent_subrecipients["subrecipients"], list)
        or len(recent_subrecipients["subrecipients"]) == 0
    )


def write_subrecipients_to_workbook(recent_subrecipients, workbook, logger):
    """
    Given an output template, in the form of a `workbook` preloaded with openpyxl,
    go through a list of `recent_subrecipients` and write information for each of them into the workbook
    """
    sheet_to_edit = workbook[WORKSHEET_NAME]
    row_to_edit = FIRST_BLANK_ROW_NUM

    for subrecipient in recent_subrecipients["subrecipients"]:
        if "subrecipientUploads" not in subrecipient:
            logger.warning(
                f"Subrecipient in recent uploads file with id {subrecipient["id"]} and name {subrecipient["Name"]} doesn't have any associated uploads, skipping in treasury report"
            )
            continue

        most_recent_upload = get_most_recent_upload(subrecipient)

        for k, v in getSubrecipientRowClass(
            version_string=most_recent_upload["version"]
        ).model_fields.items():
            output_column = v.json_schema_extra["output_column"]
            if not output_column:
                logger.error(f"No output column specified for field name {k}, skipping")
                continue

            if k in most_recent_upload["rawSubrecipient"]:
                value_to_insert = most_recent_upload["rawSubrecipient"][k]
                if value_to_insert != "null":
                    sheet_to_edit[f"{output_column}{row_to_edit}"] = value_to_insert
            else:
                # Is this helpful? Open to opinions
                logger.warning(
                    f"Did not find information in stored subrecipient data for field {k}"
                )

        row_to_edit += 1


def get_most_recent_upload(subrecipient):
    """
    Small helper method to sort subrecipientUploads for a given subrecipient by updated date,
    and return the most recent one
    """
    subrecipientUploads = subrecipient["subrecipientUploads"]
    subrecipientUploads.sort(
        key=lambda x: datetime.fromisoformat(x["updatedAt"].replace("Z", "+00:00")),
        reverse=True,
    )
    return subrecipientUploads[0]


def upload_workbook(
    workbook: Workbook,
    s3client: S3Client,
    organization: OrganizationObj,
):
    """
    Handles upload of workbook to S3, both in xlsx and csv formats
    """

    with tempfile.NamedTemporaryFile("w") as new_xlsx_file:
        workbook.save(new_xlsx_file.name)
        upload_generated_file_to_s3(
            s3client,
            os.environ["REPORTING_DATA_BUCKET_NAME"],
            get_generated_output_file_key(
                file_type=OutputFileType.XLSX,
                project="Subrecipient",
                organization=organization,
            ),
            new_xlsx_file,
        )

    with tempfile.NamedTemporaryFile("w") as new_csv_file:
        convert_xlsx_to_csv(
            new_csv_file,
            workbook,
            find_last_populated_row(
                workbook[WORKSHEET_NAME], FIRST_BLANK_ROW_NUM, FIRST_DATA_COLUMN
            ),
        )
        upload_generated_file_to_s3(
            s3client,
            os.environ["REPORTING_DATA_BUCKET_NAME"],
            get_generated_output_file_key(
                file_type=OutputFileType.CSV,
                project="Subrecipient",
                organization=organization,
            ),
            new_csv_file,
        )
